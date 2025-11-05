import requests
from bs4 import BeautifulSoup
import json
import csv
import re
from urllib.parse import urljoin, urlparse
import time
from typing import List, Dict, Optional
import pandas as pd
import os
from app.utils.progress import set_progress

# Try to import Selenium for JavaScript-heavy pages
try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False
    print("Selenium not available. Install with: pip install selenium")

class EnhancedFacultyScraper:
    def __init__(self, use_selenium=False, deep_scrape=True, max_profile_visits=None):
        self.use_selenium = use_selenium and SELENIUM_AVAILABLE
        self.deep_scrape = deep_scrape
        self.max_profile_visits = max_profile_visits
        self.visited_urls = set()
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })
        
        if self.use_selenium:
            self.setup_selenium_driver()
    
    def setup_selenium_driver(self):
        """Setup Selenium WebDriver with appropriate options."""
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1920,1080')
        
        try:
            self.driver = webdriver.Chrome(options=chrome_options)
        except Exception as e:
            print(f"Failed to setup Chrome driver: {e}")
            self.use_selenium = False
    
    def get_page_content(self, url: str) -> Optional[BeautifulSoup]:
        """Fetch and parse the webpage content."""
        if self.use_selenium:
            return self._get_content_selenium(url)
        else:
            return self._get_content_requests(url)
    
    def _get_content_requests(self, url: str) -> Optional[BeautifulSoup]:
        """Fetch content using requests."""
        try:
            response = self.session.get(url, timeout=10)
            response.raise_for_status()
            return BeautifulSoup(response.content, 'html.parser')
        except Exception as e:
            print(f"Error fetching {url} with requests: {e}")
            return None
    
    def _get_content_selenium(self, url: str) -> Optional[BeautifulSoup]:
        """Fetch content using Selenium (handles JavaScript)."""
        try:
            self.driver.get(url)
            time.sleep(4)  # Wait longer for content to load
            html = self.driver.page_source
            return BeautifulSoup(html, 'html.parser')
        except Exception as e:
            print(f"Error fetching {url} with Selenium: {e}")
            return None

    def scrape_faculty(self, url: str) -> List[Dict]:
        """Main method to scrape faculty from any university URL."""
        print(f"Scraping faculty from: {url}")
        
        soup = self.get_page_content(url)
        if not soup:
            print("Failed to fetch page content")
            return []
        
        # Detect the specific site structure and extract faculty
        faculty_list = self._extract_faculty_adaptive(soup, url)
        
        print(f"Found {len(faculty_list)} faculty members")
        
        # Show first few results for debugging
        print("\nFirst few results:")
        for i, faculty in enumerate(faculty_list[:3], 1):
            print(f"{i}. Name: {faculty.get('name', 'No name')}")
            print(f"   Title: {faculty.get('title', 'No title')}")
            print(f"   Email: {faculty.get('email', 'No email')}")
            print(f"   Profile URL: {faculty.get('profile_url', 'No URL')}")
            print()
        
        # Deep scrape individual profiles if enabled
        if self.deep_scrape and faculty_list:
            print("Starting deep scrape of individual faculty profiles...")
            faculty_list = self.deep_scrape_profiles(faculty_list)
        
        return faculty_list

    def _extract_faculty_adaptive(self, soup: BeautifulSoup, base_url: str) -> List[Dict]:
        """Adaptively extract faculty based on page structure."""
        faculty_list = []
        
        # Strategy 1: Look for faculty cards/profiles (common pattern)
        faculty_cards = self._find_faculty_cards(soup, base_url)
        faculty_list.extend(faculty_cards)
        
        # Strategy 2: Look for faculty in table rows
        table_faculty = self._extract_from_tables(soup, base_url)
        faculty_list.extend(table_faculty)
        
        # Strategy 3: Look for faculty in list items
        list_faculty = self._extract_from_lists(soup, base_url)
        faculty_list.extend(list_faculty)
        
        # Strategy 4: Look for faculty in generic containers
        container_faculty = self._extract_from_containers(soup, base_url)
        faculty_list.extend(container_faculty)
        
        # Remove duplicates and invalid entries
        faculty_list = self._clean_and_deduplicate(faculty_list)
        
        return faculty_list

    def _find_faculty_cards(self, soup: BeautifulSoup, base_url: str) -> List[Dict]:
        """Find faculty in card-like structures."""
        faculty_list = []
        
        # Common selectors for faculty cards
        card_selectors = [
            'div[class*="faculty"]',
            'div[class*="profile"]',
            'div[class*="person"]',
            'div[class*="member"]',
            'div[class*="card"]',
            'article',
            '.faculty-member',
            '.profile-card',
        ]
        
        for selector in card_selectors:
            cards = soup.select(selector)
            for card in cards:
                if self._looks_like_faculty_card(card):
                    faculty_info = self._extract_from_card(card, base_url)
                    if faculty_info and self._is_valid_faculty_info(faculty_info):
                        faculty_list.append(faculty_info)
        
        return faculty_list

    def _extract_from_tables(self, soup: BeautifulSoup, base_url: str) -> List[Dict]:
        """Extract faculty from table structures."""
        faculty_list = []
        
        tables = soup.find_all('table')
        for table in tables:
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all(['td', 'th'])
                if len(cells) >= 2:  # Need at least 2 cells for meaningful data
                    faculty_info = self._extract_from_table_row(row, cells, base_url)
                    if faculty_info and self._is_valid_faculty_info(faculty_info):
                        faculty_list.append(faculty_info)
        
        return faculty_list

    def _extract_from_lists(self, soup: BeautifulSoup, base_url: str) -> List[Dict]:
        """Extract faculty from list structures."""
        faculty_list = []
        
        # Look for lists that might contain faculty
        lists = soup.find_all(['ul', 'ol'])
        for ul in lists:
            items = ul.find_all('li')
            for item in items:
                if self._looks_like_faculty_item(item):
                    faculty_info = self._extract_from_list_item(item, base_url)
                    if faculty_info and self._is_valid_faculty_info(faculty_info):
                        faculty_list.append(faculty_info)
        
        return faculty_list

    def _extract_from_containers(self, soup: BeautifulSoup, base_url: str) -> List[Dict]:
        """Extract faculty from generic div containers."""
        faculty_list = []
        
        # Look for divs that might contain individual faculty
        divs = soup.find_all('div')
        for div in divs:
            # Skip if it's too nested or too large (likely container, not individual faculty)
            if len(div.find_all('div')) > 10 or len(div.get_text()) > 2000:
                continue
                
            if self._looks_like_faculty_container(div):
                faculty_info = self._extract_from_generic_container(div, base_url)
                if faculty_info and self._is_valid_faculty_info(faculty_info):
                    faculty_list.append(faculty_info)
        
        return faculty_list

    def _looks_like_faculty_card(self, element) -> bool:
        """Check if element looks like a faculty card/profile."""
        text = element.get_text().lower()
        
        # Must have reasonable amount of text
        if len(text) < 30 or len(text) > 1000:
            return False
        
        # Look for faculty indicators
        faculty_indicators = ['professor', 'dr.', 'ph.d', 'assistant', 'associate', 'chair', 'director']
        has_title = any(indicator in text for indicator in faculty_indicators)
        
        # Look for email
        has_email = '@' in text
        
        # Look for name pattern
        has_name = bool(re.search(r'\b[A-Z][a-z]+\s+[A-Z][a-z]+\b', element.get_text()))
        
        return has_name and (has_title or has_email)

    def _looks_like_faculty_item(self, element) -> bool:
        """Check if list item looks like faculty info."""
        text = element.get_text().strip()
        
        # Should have reasonable length
        if len(text) < 20 or len(text) > 500:
            return False
        
        # Look for name + title/email pattern
        has_name = bool(re.search(r'\b[A-Z][a-z]+\s+[A-Z][a-z]+\b', text))
        has_email = '@' in text
        has_title = any(word in text.lower() for word in ['professor', 'dr.', 'chair', 'director'])
        
        return has_name and (has_email or has_title)

    def _looks_like_faculty_container(self, element) -> bool:
        """Check if container looks like it contains faculty info."""
        text = element.get_text().strip()
        
        if len(text) < 40 or len(text) > 800:
            return False
        
        # Look for combinations that suggest faculty
        has_name = bool(re.search(r'\b[A-Z][a-z]+\s+[A-Z][a-z]+\b', text))
        has_email = '@' in text and '.edu' in text
        has_title = any(word in text.lower() for word in ['professor', 'dr.', 'assistant', 'associate'])
        has_contact = bool(re.search(r'\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}', text))  # Phone pattern
        
        # Need name plus at least one other indicator
        return has_name and (has_email or has_title or has_contact)

    def _extract_from_card(self, card, base_url: str) -> Dict:
        """Extract faculty info from a card-like element."""
        faculty_info = {}
        
        # Extract name - prioritize headings and links
        name = self._extract_name_improved(card)
        if name:
            faculty_info['name'] = name
        
        # Extract other info
        faculty_info.update(self._extract_contact_info(card, base_url))
        
        return faculty_info

    def _extract_from_table_row(self, row, cells, base_url: str) -> Dict:
        """Extract faculty info from table row."""
        faculty_info = {}
        
        # Combine all cell text for analysis
        combined_text = ' '.join(cell.get_text(strip=True) for cell in cells)
        
        # Extract name from the row
        name = self._extract_name_from_text(combined_text, row)
        if name:
            faculty_info['name'] = name
            faculty_info.update(self._extract_contact_info(row, base_url))
        
        return faculty_info

    def _extract_from_list_item(self, item, base_url: str) -> Dict:
        """Extract faculty info from list item."""
        faculty_info = {}
        
        name = self._extract_name_improved(item)
        if name:
            faculty_info['name'] = name
            faculty_info.update(self._extract_contact_info(item, base_url))
        
        return faculty_info

    def _extract_from_generic_container(self, container, base_url: str) -> Dict:
        """Extract faculty info from generic container."""
        faculty_info = {}
        
        name = self._extract_name_improved(container)
        if name:
            faculty_info['name'] = name
            faculty_info.update(self._extract_contact_info(container, base_url))
        
        return faculty_info

    def _extract_name_improved(self, element) -> Optional[str]:
        """Improved name extraction that avoids titles."""
        # Strategy 1: Look for links that contain names (often faculty names are linked)
        links = element.find_all('a', href=True)
        for link in links:
            href = link.get('href', '')
            text = link.get_text(strip=True)
            
            # Skip mailto links when looking for names
            if href.startswith('mailto:'):
                continue
            
            # Check if link text looks like a name
            if self._is_person_name(text):
                return text
        
        # Strategy 2: Look in headings
        for tag in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
            heading = element.find(tag)
            if heading:
                text = heading.get_text(strip=True)
                if self._is_person_name(text):
                    return text
        
        # Strategy 3: Look in strong/bold tags
        for strong in element.find_all(['strong', 'b']):
            text = strong.get_text(strip=True)
            if self._is_person_name(text):
                return text
        
        # Strategy 4: Use regex to find name patterns, but be more selective
        text = element.get_text()
        name = self._extract_name_from_text(text, element)
        if name:
            return name
        
        return None

    def _extract_name_from_text(self, text: str, element) -> Optional[str]:
        """Extract name from text using improved patterns."""
        # Pattern 1: Dr. FirstName LastName
        match = re.search(r'\bDr\.?\s+([A-Z][a-z]+\s+[A-Z][a-z]+)\b', text)
        if match:
            name = match.group(1)
            if self._is_person_name(name):
                return name
        
        # Pattern 2: FirstName LastName, Title (but extract just the name part)
        match = re.search(r'\b([A-Z][a-z]+\s+[A-Z][a-z]+)(?:\s*,\s*(?:Professor|Ph\.?D\.?|Chair))', text)
        if match:
            name = match.group(1)
            if self._is_person_name(name):
                return name
        
        # Pattern 3: Simple FirstName LastName (but validate it's not a title)
        matches = re.findall(r'\b[A-Z][a-z]+\s+[A-Z][a-z]+\b', text)
        for match in matches:
            if self._is_person_name(match):
                # Additional check: make sure it's not embedded in a title
                context = text[max(0, text.find(match) - 20):text.find(match) + len(match) + 20]
                if not any(title_word in context.lower() for title_word in ['professor of', 'department of', 'chair of']):
                    return match
        
        return None

    def _is_person_name(self, text: str) -> bool:
        """Improved check if text is a person's name."""
        if not text or len(text) > 50:
            return False
        
        # Clean up text
        text = text.strip()
        words = text.split()
        
        # Should have 2-4 words
        if len(words) < 2 or len(words) > 4:
            return False
        
        # Should not contain numbers
        if any(char.isdigit() for char in text):
            return False
        
        # Each word should start with capital letter
        if not all(word[0].isupper() for word in words if word):
            return False
        
        # Should not be common academic titles or department names
        title_words = [
            'professor', 'assistant', 'associate', 'chair', 'director', 'dean',
            'department', 'college', 'school', 'university', 'research', 'computer',
            'science', 'engineering', 'mathematics', 'physics', 'biology', 'chemistry'
        ]
        
        text_lower = text.lower()
        if any(title_word in text_lower for title_word in title_words):
            return False
        
        # Should not be common non-name phrases
        non_names = ['more info', 'contact us', 'learn more', 'click here', 'read more']
        if any(non_name in text_lower for non_name in non_names):
            return False
        
        return True

    def _extract_contact_info(self, element, base_url: str) -> Dict:
        """Extract contact information from element."""
        info = {}
        text = element.get_text()
        
        # Extract title
        title = self._extract_title_improved(text)
        if title:
            info['title'] = title
        
        # Extract email
        email = self._extract_email(element)
        if email:
            info['email'] = email
        
        # Extract phone
        phone_match = re.search(r'\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}', text)
        if phone_match:
            info['phone'] = phone_match.group(0).strip()
        
        # Extract office
        office_patterns = [
            r'(?:Office|Room|Building):?\s*([A-Z]{1,4}[-\s]*\d+[A-Z]?)',
            r'(?:Office|Room):?\s*(\d+[A-Z]?\s+[A-Z][a-z]+)',
        ]
        for pattern in office_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                info['office'] = match.group(1).strip()
                break
        
        # Extract profile URL
        profile_url = self._extract_profile_url_improved(element, base_url)
        if profile_url:
            info['profile_url'] = profile_url
        
        # Extract image
        img = element.find('img')
        if img and img.get('src'):
            info['image_url'] = urljoin(base_url, img['src'])
        
        return info

    def _extract_title_improved(self, text: str) -> Optional[str]:
        """Extract academic title with better precision."""
        title_patterns = [
            r'((?:Assistant|Associate|Full)\s+Professor(?:\s+of\s+[A-Za-z\s]+)?)',
            r'(Professor(?:\s+of\s+[A-Za-z\s]+)?)',
            r'(Chair(?:\s+of\s+[A-Za-z\s]+)?)',
            r'(Director(?:\s+of\s+[A-Za-z\s]+)?)',
            r'(Lecturer)',
            r'(Instructor)',
            r'(Research\s+(?:Professor|Scientist))',
            r'(Clinical\s+Professor)',
            r'(Emeritus\s+Professor)',
        ]
        
        for pattern in title_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                title = match.group(1).strip()
                # Make sure it's not too long (probably picked up too much context)
                if len(title) < 60:
                    return title
        
        return None

    def _extract_email(self, element) -> Optional[str]:
        """Extract email with improved methods."""
        # Try mailto links first
        mailto_links = element.find_all('a', href=re.compile(r'^mailto:'))
        if mailto_links:
            return mailto_links[0]['href'].replace('mailto:', '').strip()
        
        # Try text extraction
        text = element.get_text()
        email_match = re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', text)
        return email_match.group(0) if email_match else None

    def _extract_profile_url_improved(self, element, base_url: str) -> Optional[str]:
        """Extract profile URL with better logic."""
        links = element.find_all('a', href=True)
        
        # First priority: links with profile-related text
        for link in links:
            text = link.get_text().lower()
            if any(word in text for word in ['profile', 'bio', 'cv', 'more', 'details', 'homepage']):
                href = link['href']
                if not href.startswith('mailto:'):
                    return urljoin(base_url, href)
        
        # Second priority: links containing the person's name
        name_from_element = self._extract_name_improved(element)
        if name_from_element:
            for link in links:
                link_text = link.get_text()
                if name_from_element.lower() in link_text.lower():
                    href = link['href']
                    if not href.startswith('mailto:') and not href.startswith('#'):
                        return urljoin(base_url, href)
        
        # Last priority: any non-mailto link
        for link in links:
            href = link['href']
            if href and not href.startswith('mailto:') and not href.startswith('#'):
                return urljoin(base_url, href)
        
        return None

    def _is_valid_faculty_info(self, faculty_info: Dict) -> bool:
        """Check if extracted faculty info is valid."""
        # Must have a name
        if not faculty_info.get('name'):
            return False
        
        # Name should not be a title
        name = faculty_info['name'].lower()
        if any(word in name for word in ['professor', 'chair', 'director', 'department']):
            return False
        
        # Should have at least one other piece of information
        other_fields = ['title', 'email', 'phone', 'office', 'profile_url']
        return any(faculty_info.get(field) for field in other_fields)

    def _clean_and_deduplicate(self, faculty_list: List[Dict]) -> List[Dict]:
        """Clean and remove duplicates from faculty list."""
        seen = set()
        cleaned_list = []
        
        for faculty in faculty_list:
            name = faculty.get('name', '').strip()
            email = faculty.get('email', '').strip()
            
            # Skip if no name
            if not name:
                continue
            
            # Create identifier for deduplication
            identifier = (name.lower(), email.lower())
            
            if identifier not in seen:
                seen.add(identifier)
                cleaned_list.append(faculty)
        
        return cleaned_list

    def deep_scrape_profiles(self, faculty_list: List[Dict]) -> List[Dict]:
        """Visit individual profiles for detailed research interests."""
        enhanced_faculty = []
        max_visits = self.max_profile_visits or len(faculty_list)
        
        print(f"Will visit {min(max_visits, len(faculty_list))} profile pages for research interests")
        
        for i, faculty in enumerate(faculty_list):
            if i >= max_visits:
                enhanced_faculty.extend(faculty_list[i:])
                break
            
            percent = int(((i + 1) / max_visits) * 100)
            set_progress("scrape", percent)

            profile_url = faculty.get('profile_url')
            name = faculty.get('name', f'Faculty {i+1}')
            
            if profile_url:
                print(f"Scraping profile {i+1}/{min(max_visits, len(faculty_list))}: {name}")
                
                try:
                    detailed_info = self.scrape_individual_profile(profile_url)
                    enhanced_faculty_member = {**faculty, **detailed_info}
                    enhanced_faculty.append(enhanced_faculty_member)
                    time.sleep(1.5)  # Be respectful with delays
                except Exception as e:
                    print(f"   Error: {e}")
                    enhanced_faculty.append(faculty)
            else:
                print(f"Skipping {name} - no profile URL found")
                enhanced_faculty.append(faculty)
        
        return enhanced_faculty

    def scrape_individual_profile(self, profile_url: str) -> Dict:
        """Scrape detailed info from individual profile page."""
        soup = self.get_page_content(profile_url)
        if not soup:
            return {}
        
        detailed_info = {}
        
        # Extract research interests with multiple strategies
        research = self._extract_research_comprehensive(soup)
        if research:
            detailed_info['research_interests'] = research
            print(f"   Found research: {research[:80]}...")
        
        # Extract education
        education = self._extract_education_info(soup)
        if education:
            detailed_info['education'] = education
        
        # Extract biography
        bio = self._extract_biography_info(soup)
        if bio:
            detailed_info['biography'] = bio[:800]  # Limit length
        
        return detailed_info

    def _extract_research_comprehensive(self, soup: BeautifulSoup) -> Optional[str]:
        """Extract research interests using comprehensive strategies."""
        research_sections = []
        
        # Strategy 1: Look for dedicated research sections by headings
        research_keywords = ['research', 'interests', 'areas', 'focus', 'specialization', 'expertise']
        headings = soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6'])
        
        for heading in headings:
            heading_text = heading.get_text().lower()
            if any(keyword in heading_text for keyword in research_keywords):
                # Extract content following this heading
                content_parts = self._extract_content_after_heading(heading)
                if content_parts:
                    research_sections.extend(content_parts)
        
        # Strategy 2: Look for research-related paragraphs
        paragraphs = soup.find_all('p')
        research_context_words = [
            'research', 'studying', 'investigating', 'focus', 'specializ',
            'work on', 'interests include', 'projects', 'current work'
        ]
        
        for p in paragraphs:
            text = p.get_text()
            if any(word in text.lower() for word in research_context_words) and len(text) > 80:
                research_sections.append(text.strip())
        
        # Strategy 3: Look for research in lists
        lists = soup.find_all(['ul', 'ol'])
        for ul in lists:
            # Check if previous content suggests this list is about research
            prev_elements = []
            prev = ul.previous_sibling
            for _ in range(3):  # Check previous 3 elements
                if prev and hasattr(prev, 'get_text'):
                    prev_elements.append(prev.get_text().lower())
                    prev = prev.previous_sibling
                elif prev:
                    prev = prev.previous_sibling
                else:
                    break
            
            prev_context = ' '.join(prev_elements)
            if any(keyword in prev_context for keyword in research_keywords):
                items = [li.get_text(strip=True) for li in ul.find_all('li')]
                if items and len(' '.join(items)) > 50:  # Substantial content
                    research_sections.append(' | '.join(items))
        
        # Strategy 4: Look for research in divs with research-related classes or IDs
        research_divs = soup.find_all('div', {'class': re.compile(r'research|interest', re.I)})
        research_divs.extend(soup.find_all('div', {'id': re.compile(r'research|interest', re.I)}))
        
        for div in research_divs:
            text = div.get_text(strip=True)
            if len(text) > 50:
                research_sections.append(text)
        
        # Combine and clean research sections
        if research_sections:
            # Remove duplicates while preserving order
            unique_sections = []
            seen_content = set()
            
            for section in research_sections:
                # Create a normalized version for comparison
                normalized = re.sub(r'\s+', ' ', section.lower().strip())
                if len(normalized) > 30 and normalized not in seen_content:
                    seen_content.add(normalized)
                    unique_sections.append(section.strip())
            
            if unique_sections:
                combined = ' | '.join(unique_sections[:4])  # Limit to top 4 sections
                # Clean up the text
                combined = re.sub(r'\s+', ' ', combined)
                return combined[:2000]  # Limit total length
        
        return None

    def _extract_content_after_heading(self, heading) -> List[str]:
        """Extract content that follows a heading until the next heading."""
        content_parts = []
        current = heading.next_sibling
        
        while current and len(content_parts) < 5:  # Limit to avoid too much content
            if hasattr(current, 'name'):
                # Stop if we hit another heading
                if current.name in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
                    break
                
                # Extract text from paragraphs, divs, and lists
                if current.name in ['p', 'div', 'ul', 'ol']:
                    text = current.get_text(strip=True)
                    if text and len(text) > 30:  # Meaningful content
                        content_parts.append(text)
            
            current = current.next_sibling
        
        return content_parts

    def _extract_education_info(self, soup: BeautifulSoup) -> Optional[str]:
        """Extract education information from profile page."""
        education_sections = []
        education_keywords = ['education', 'degrees', 'phd', 'ph.d', 'master', 'bachelor', 'university', 'degree']
        
        # Look for education-related headings
        headings = soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6'])
        for heading in headings:
            heading_text = heading.get_text().lower()
            if any(keyword in heading_text for keyword in education_keywords):
                content = self._extract_content_after_heading(heading)
                if content:
                    education_sections.extend(content)
        
        # Look for education in paragraphs
        paragraphs = soup.find_all('p')
        for p in paragraphs:
            text = p.get_text()
            if any(keyword in text.lower() for keyword in education_keywords) and len(text) > 40:
                education_sections.append(text.strip())
        
        if education_sections:
            return ' | '.join(education_sections[:3])[:1000]  # Limit length
        
        return None

    def _extract_biography_info(self, soup: BeautifulSoup) -> Optional[str]:
        """Extract biography information from profile page."""
        bio_sections = []
        bio_keywords = ['biography', 'about', 'bio', 'overview', 'background']
        
        # Look for bio-related headings
        headings = soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6'])
        for heading in headings:
            heading_text = heading.get_text().lower()
            if any(keyword in heading_text for keyword in bio_keywords):
                content = self._extract_content_after_heading(heading)
                if content:
                    bio_sections.extend(content)
        
        if bio_sections:
            return ' '.join(bio_sections[:2])[:1200]  # Limit length
        
        return None

    # def save_to_excel(self, faculty_list: List[Dict], filename: str = 'faculty.xlsx'):
    #     """Save faculty data to Excel with formatting."""
    #     if not faculty_list:
    #         print("No faculty data to save")
    #         return
        
    #     df = pd.DataFrame(faculty_list)
        
    #     # Reorder columns for better presentation
    #     preferred_order = [
    #         'name', 'title', 'email', 'phone', 'office', 
    #         'research_interests', 'education', 'biography', 
    #         'profile_url', 'image_url'
    #     ]
        
    #     available_cols = [col for col in preferred_order if col in df.columns]
    #     other_cols = [col for col in df.columns if col not in preferred_order]
    #     final_columns = available_cols + other_cols
        
    #     if final_columns:
    #         df = df[final_columns]
        
    #     # Save with formatting
    #     with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    #         df.to_excel(writer, sheet_name='Faculty Directory', index=False)
            
    #         worksheet = writer.sheets['Faculty Directory']
            
    #         # Set column widths based on content type
    #         column_widths = {
    #             'name': 25,
    #             'title': 35,
    #             'email': 30,
    #             'phone': 15,
    #             'office': 15,
    #             'research_interests': 70,
    #             'education': 45,
    #             'biography': 50,
    #             'profile_url': 40,
    #             'image_url': 30
    #         }
            
    #         for i, column in enumerate(df.columns, 1):
    #             col_letter = chr(64 + i) if i <= 26 else chr(64 + i // 26) + chr(64 + i % 26)
    #             width = column_widths.get(column, 20)
    #             worksheet.column_dimensions[col_letter].width = width
            
    #         # Add header formatting
    #         from openpyxl.styles import Font, PatternFill, Alignment
            
    #         header_font = Font(bold=True, color='FFFFFF')
    #         header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    #         header_alignment = Alignment(horizontal='center', vertical='center')
            
    #         for col in range(1, len(df.columns) + 1):
    #             cell = worksheet.cell(row=1, column=col)
    #             cell.font = header_font
    #             cell.fill = header_fill
    #             cell.alignment = header_alignment
        
    #     print(f"Saved to {filename}")
    #     print(f"Total faculty: {len(df)}")
        
    #     # Show statistics
    #     stats = {}
    #     for field in ['research_interests', 'education', 'biography', 'email', 'phone']:
    #         if field in df.columns:
    #             count = len([f for f in faculty_list if f.get(field)])
    #             if count > 0:
    #                 stats[field] = count
        
    #     print("\nData completeness:")
    #     for field, count in stats.items():
    #         print(f"  {field}: {count}/{len(df)} ({count/len(df)*100:.1f}%)")

    # def __del__(self):
    #     if hasattr(self, 'driver'):
    #         try:
    #             self.driver.quit()
    #         except:
    #             pass
    def save_to_excel(self, faculty_list: List[Dict], filename: str = 'faculty.xlsx'):
        """Save faculty data to Excel with formatting into faculty_documents_deep folder."""
        if not faculty_list:
            print("No faculty data to save")
            return

        # Ensure the output directory exists
        output_dir = os.path.join(os.getcwd(), "faculty_documents_deep")
        os.makedirs(output_dir, exist_ok=True)

        # Build full path
        file_path = os.path.join(output_dir, filename)

        df = pd.DataFrame(faculty_list)

        # Reorder columns for better presentation
        preferred_order = [
            'name', 'title', 'email', 'phone', 'office',
            'research_interests', 'education', 'biography',
            'profile_url', 'image_url'
        ]

        available_cols = [col for col in preferred_order if col in df.columns]
        other_cols = [col for col in df.columns if col not in preferred_order]
        final_columns = available_cols + other_cols

        if final_columns:
            df = df[final_columns]

        # Save with formatting
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Faculty Directory', index=False)
            worksheet = writer.sheets['Faculty Directory']

            # Set column widths
            column_widths = {
                'name': 25, 'title': 35, 'email': 30, 'phone': 15,
                'office': 15, 'research_interests': 70, 'education': 45,
                'biography': 50, 'profile_url': 40, 'image_url': 30
            }

            for i, column in enumerate(df.columns, 1):
                col_letter = chr(64 + i) if i <= 26 else chr(64 + i // 26) + chr(64 + i % 26)
                width = column_widths.get(column, 20)
                worksheet.column_dimensions[col_letter].width = width

            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')

            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

        print(f"Saved to {file_path}")
        print(f"Total faculty: {len(df)}")

        # Stats
        stats = {}
        for field in ['research_interests', 'education', 'biography', 'email', 'phone']:
            if field in df.columns:
                count = len([f for f in faculty_list if f.get(field)])
                if count > 0:
                    stats[field] = count

        print("\nData completeness:")
        for field, count in stats.items():
            print(f"  {field}: {count}/{len(df)} ({count/len(df)*100:.1f}%)")



# Usage for Illinois site
if __name__ == "__main__":
    print(f"Files will be saved in: {os.getcwd()}")
    
    scraper = EnhancedFacultyScraper(
        use_selenium=True,
        deep_scrape=True,
        max_profile_visits=None  # Test with 10 first
    )
    
    faculty_data = scraper.scrape_faculty("https://www.depts.ttu.edu/cs/faculty/")
    
    if faculty_data:
        scraper.save_to_excel(faculty_data, 'illinois_faculty_detailed.xlsx')
        print(f"\nExcel file saved in: {os.path.join(os.getcwd(), 'faculty_documents_deep', 'illinois_faculty_detailed.xlsx')}")

    else:
        print("No faculty data found")