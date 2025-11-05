# import requests
# from bs4 import BeautifulSoup
# import json
# import csv
# import re
# from urllib.parse import urljoin, urlparse
# import time
# from typing import List, Dict, Optional

# # Try to import Selenium for JavaScript-heavy pages
# try:
#     from selenium import webdriver
#     from selenium.webdriver.common.by import By
#     from selenium.webdriver.chrome.options import Options
#     from selenium.webdriver.support.ui import WebDriverWait
#     from selenium.webdriver.support import expected_conditions as EC
#     SELENIUM_AVAILABLE = True
# except ImportError:
#     SELENIUM_AVAILABLE = False
#     print("Selenium not available. Install with: pip install selenium")
#     print("Also download ChromeDriver from: https://chromedriver.chromium.org/")

# class FacultyScraper:
#     def __init__(self, use_selenium=False):
#         self.use_selenium = use_selenium and SELENIUM_AVAILABLE
#         self.session = requests.Session()
#         self.session.headers.update({
#             'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
#         })
        
#         if self.use_selenium:
#             self.setup_selenium_driver()
    
#     def setup_selenium_driver(self):
#         """Setup Selenium WebDriver with appropriate options."""
#         chrome_options = Options()
#         chrome_options.add_argument('--headless')  # Run in background
#         chrome_options.add_argument('--no-sandbox')
#         chrome_options.add_argument('--disable-dev-shm-usage')
#         chrome_options.add_argument('--disable-gpu')
#         chrome_options.add_argument('--window-size=1920,1080')
        
#         try:
#             self.driver = webdriver.Chrome(options=chrome_options)
#         except Exception as e:
#             print(f"Failed to setup Chrome driver: {e}")
#             print("Make sure ChromeDriver is installed and in PATH")
#             self.use_selenium = False
    
#     def get_page_content(self, url: str) -> Optional[BeautifulSoup]:
#         """Fetch and parse the webpage content."""
#         if self.use_selenium:
#             return self._get_content_selenium(url)
#         else:
#             return self._get_content_requests(url)
    
#     def _get_content_requests(self, url: str) -> Optional[BeautifulSoup]:
#         """Fetch content using requests."""
#         try:
#             response = self.session.get(url, timeout=10)
#             response.raise_for_status()
#             return BeautifulSoup(response.content, 'html.parser')
#         except Exception as e:
#             print(f"Error fetching {url} with requests: {e}")
#             return None
    
#     def _get_content_selenium(self, url: str) -> Optional[BeautifulSoup]:
#         """Fetch content using Selenium (handles JavaScript)."""
#         try:
#             self.driver.get(url)
            
#             # Wait for potential faculty content to load
#             try:
#                 WebDriverWait(self.driver, 10).until(
#                     lambda driver: len(driver.find_elements(By.CSS_SELECTOR, 
#                         '[class*="faculty"], [class*="profile"], [class*="person"], h2, h3')) > 0
#                 )
#             except:
#                 # Timeout is okay, just continue with what's loaded
#                 pass
            
#             # Additional wait for any AJAX content
#             time.sleep(3)
            
#             html = self.driver.page_source
#             return BeautifulSoup(html, 'html.parser')
#         except Exception as e:
#             print(f"Error fetching {url} with Selenium: {e}")
#             return None
    
#     def debug_page_structure(self, soup: BeautifulSoup, url: str):
#         """Debug function to understand page structure."""
#         print(f"\n=== DEBUG: Page Structure Analysis for {url} ===")
        
#         # Check for common faculty-related elements
#         selectors_to_check = {
#             'Faculty containers': '[class*="faculty"]',
#             'Profile containers': '[class*="profile"]', 
#             'Person containers': '[class*="person"]',
#             'Card containers': '[class*="card"]',
#             'All headings': 'h1, h2, h3, h4, h5, h6',
#             'Links': 'a[href]',
#             'Images': 'img',
#             'Email links': 'a[href*="mailto"]',
#             'Divs with IDs': 'div[id]',
#             'Divs with classes': 'div[class]',
#         }
        
#         for name, selector in selectors_to_check.items():
#             elements = soup.select(selector)
#             print(f"{name}: {len(elements)} found")
            
#             # Show first few examples
#             for i, elem in enumerate(elements[:3]):
#                 if name == 'All headings':
#                     print(f"  {i+1}. {elem.name}: {elem.get_text()[:100].strip()}")
#                 elif name == 'Links':
#                     print(f"  {i+1}. {elem.get('href', '')}: {elem.get_text()[:50].strip()}")
#                 elif name == 'Images':
#                     print(f"  {i+1}. {elem.get('src', '')} - alt: {elem.get('alt', '')}")
#                 else:
#                     classes = ' '.join(elem.get('class', []))
#                     text_preview = elem.get_text()[:100].strip().replace('\n', ' ')
#                     print(f"  {i+1}. class='{classes}': {text_preview}")
        
#         # Check for JavaScript-loaded content indicators
#         scripts = soup.find_all('script')
#         js_indicators = ['ajax', 'fetch', 'XMLHttpRequest', 'faculty', 'profile']
#         js_found = []
#         for script in scripts:
#             script_text = script.get_text().lower() if script.string else ''
#             for indicator in js_indicators:
#                 if indicator in script_text:
#                     js_found.append(indicator)
        
#         if js_found:
#             print(f"JavaScript indicators found: {set(js_found)}")
#             print("Page likely loads content dynamically - consider using Selenium")
        
#         print("=== END DEBUG ===\n")
    
#     def extract_faculty_info(self, soup: BeautifulSoup, base_url: str, debug=True) -> List[Dict]:
#         """Extract faculty information using multiple strategies."""
#         if debug:
#             self.debug_page_structure(soup, base_url)
        
#         faculty_list = []
        
#         # Strategy 1: Texas A&M specific patterns
#         tamu_faculty = self._extract_tamu_specific(soup, base_url)
#         faculty_list.extend(tamu_faculty)
        
#         # Strategy 2: Common faculty card/profile containers
#         faculty_containers = self._find_faculty_containers(soup)
        
#         for container in faculty_containers:
#             faculty_info = self._extract_from_container(container, base_url)
#             if faculty_info and faculty_info not in faculty_list:
#                 faculty_list.append(faculty_info)
        
#         # Strategy 3: Extract from headings and surrounding content
#         heading_faculty = self._extract_from_headings(soup, base_url)
#         faculty_list.extend(heading_faculty)
        
#         # Strategy 4: Look for structured data (JSON-LD, microdata)
#         structured_data = self._extract_structured_data(soup)
#         faculty_list.extend(structured_data)
        
#         # Remove duplicates based on name or email
#         faculty_list = self._remove_duplicates(faculty_list)
        
#         return faculty_list
    
#     def _extract_tamu_specific(self, soup: BeautifulSoup, base_url: str) -> List[Dict]:
#         """Texas A&M specific extraction patterns."""
#         faculty_list = []
        
#         # Look for TAMU-specific patterns
#         tamu_selectors = [
#             '.faculty-list .faculty-item',
#             '.profile-card',
#             '.person-card',
#             '[data-faculty]',
#             '.directory-item',
#             '.profile-listing',
#         ]
        
#         for selector in tamu_selectors:
#             elements = soup.select(selector)
#             for element in elements:
#                 faculty_info = self._extract_from_container(element, base_url)
#                 if faculty_info:
#                     faculty_list.append(faculty_info)
        
#         return faculty_list
    
#     def _extract_from_headings(self, soup: BeautifulSoup, base_url: str) -> List[Dict]:
#         """Extract faculty info from headings and surrounding content."""
#         faculty_list = []
        
#         # Look for headings that might contain names
#         headings = soup.find_all(['h2', 'h3', 'h4', 'h5'])
        
#         for heading in headings:
#             text = heading.get_text().strip()
            
#             # Check if heading looks like a person's name
#             if self._looks_like_person_name(text):
#                 faculty_info = {'name': text}
                
#                 # Look for additional info in siblings or parent
#                 parent = heading.parent
#                 if parent:
#                     faculty_info.update(self._extract_from_container(parent, base_url))
                
#                 # Look in next siblings
#                 for sibling in heading.find_next_siblings():
#                     sibling_text = sibling.get_text()
                    
#                     # Extract email
#                     email = self._extract_email_from_text(sibling_text)
#                     if email and 'email' not in faculty_info:
#                         faculty_info['email'] = email
                    
#                     # Extract title
#                     if not faculty_info.get('title'):
#                         title = self._extract_title_from_text(sibling_text)
#                         if title:
#                             faculty_info['title'] = title
                    
#                     # Stop after a few siblings or if we hit another heading
#                     if sibling.name in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
#                         break
                
#                 if len(faculty_info) > 1:  # More than just name
#                     faculty_list.append(faculty_info)
        
#         return faculty_list
    
#     def _looks_like_person_name(self, text: str) -> bool:
#         """Check if text looks like a person's name."""
#         words = text.split()
        
#         # Should have 2-4 words
#         if len(words) < 2 or len(words) > 4:
#             return False
        
#         # Shouldn't contain numbers
#         if any(char.isdigit() for char in text):
#             return False
        
#         # Shouldn't be too long
#         if len(text) > 50:
#             return False
        
#         # Should start with capital letters
#         if not all(word[0].isupper() for word in words if word):
#             return False
        
#         # Common non-name patterns
#         exclude_patterns = ['about', 'contact', 'research', 'education', 'department']
#         if any(pattern in text.lower() for pattern in exclude_patterns):
#             return False
        
#         return True
    
#     def _extract_email_from_text(self, text: str) -> Optional[str]:
#         """Extract email from text."""
#         email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
#         match = re.search(email_pattern, text)
#         return match.group(0) if match else None
    
#     def _extract_title_from_text(self, text: str) -> Optional[str]:
#         """Extract academic title from text."""
#         title_patterns = [
#             r'(Professor|Prof\.)\s*[A-Za-z\s]*',
#             r'(Assistant|Associate|Full)\s+Professor',
#             r'(Chair|Director)\s+of\s+[A-Za-z\s]+',
#             r'(Dr\.|PhD|Ph\.D\.)',
#             r'(Lecturer|Instructor)',
#         ]
        
#         for pattern in title_patterns:
#             match = re.search(pattern, text, re.IGNORECASE)
#             if match:
#                 return match.group(0).strip()
        
#         return None
    
#     def _find_faculty_containers(self, soup: BeautifulSoup) -> List:
#         """Find potential faculty containers using various selectors."""
#         selectors = [
#             # Common class patterns for faculty profiles
#             '[class*="faculty"]',
#             '[class*="profile"]',
#             '[class*="person"]',
#             '[class*="staff"]',
#             '[class*="member"]',
#             '[class*="bio"]',
#             '[class*="card"]',
#             '[class*="directory"]',
#             # Common ID patterns
#             '[id*="faculty"]',
#             '[id*="profile"]',
#             # Semantic elements
#             'article',
#             '.row > div',
#             '.col > div',
#             # List items that might contain faculty
#             'li[class*="faculty"]',
#             'li[class*="profile"]',
#             'li[class*="person"]',
#         ]
        
#         containers = []
#         for selector in selectors:
#             elements = soup.select(selector)
#             for element in elements:
#                 # Check if element likely contains faculty info
#                 if self._looks_like_faculty_container(element):
#                     containers.append(element)
        
#         return containers
    
#     def _looks_like_faculty_container(self, element) -> bool:
#         """Determine if an element likely contains faculty information."""
#         text = element.get_text().lower()
        
#         # Look for academic titles
#         academic_titles = ['professor', 'dr.', 'ph.d', 'assistant', 'associate', 'chair', 'director', 'lecturer']
#         has_title = any(title in text for title in academic_titles)
        
#         # Look for email patterns
#         has_email = '@' in text and ('edu' in text or 'ac.' in text)
        
#         # Look for common faculty page indicators
#         faculty_indicators = ['research', 'education', 'cv', 'publications', 'office', 'phone', 'interests']
#         has_indicators = sum(1 for indicator in faculty_indicators if indicator in text) >= 1
        
#         # Check for images (faculty photos)
#         has_image = element.find('img') is not None
        
#         # Should have reasonable amount of text
#         has_content = 20 <= len(text.strip()) <= 2000
        
#         # Look for person name patterns
#         has_name = bool(re.search(r'\b[A-Z][a-z]+\s+[A-Z][a-z]+\b', element.get_text()))
        
#         return (has_title or has_email or has_name) and has_content
    
#     def _extract_from_container(self, container, base_url: str) -> Dict:
#         """Extract faculty information from a container element."""
#         faculty_info = {}
        
#         # Extract name (usually in headings or strong/bold text)
#         name = self._extract_name(container)
#         if name:
#             faculty_info['name'] = name
        
#         # Extract title/position
#         title = self._extract_title(container)
#         if title:
#             faculty_info['title'] = title
        
#         # Extract email
#         email = self._extract_email(container)
#         if email:
#             faculty_info['email'] = email
        
#         # Extract phone
#         phone = self._extract_phone(container)
#         if phone:
#             faculty_info['phone'] = phone
        
#         # Extract office/location
#         office = self._extract_office(container)
#         if office:
#             faculty_info['office'] = office
        
#         # Extract profile URL
#         profile_url = self._extract_profile_url(container, base_url)
#         if profile_url:
#             faculty_info['profile_url'] = profile_url
        
#         # Extract image URL
#         image_url = self._extract_image_url(container, base_url)
#         if image_url:
#             faculty_info['image_url'] = image_url
        
#         # Extract research areas/interests
#         research = self._extract_research_areas(container)
#         if research:
#             faculty_info['research_areas'] = research
        
#         return faculty_info if faculty_info else {}
    
#     def _extract_name(self, container) -> Optional[str]:
#         """Extract faculty name from container."""
#         # Try headings first
#         for tag in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
#             heading = container.find(tag)
#             if heading:
#                 name = heading.get_text(strip=True)
#                 if self._looks_like_person_name(name):
#                     return name
        
#         # Try strong/bold text
#         strong_tags = container.find_all(['strong', 'b'])
#         for strong in strong_tags:
#             text = strong.get_text(strip=True)
#             if self._looks_like_person_name(text):
#                 return text
        
#         # Try links that might be names
#         links = container.find_all('a')
#         for link in links:
#             text = link.get_text(strip=True)
#             if self._looks_like_person_name(text):
#                 return text
        
#         # Try to find name patterns in the text
#         text = container.get_text()
#         name_pattern = r'\b[A-Z][a-z]+(?:\s+[A-Z][a-z]*\.?)?\s+[A-Z][a-z]+\b'
#         match = re.search(name_pattern, text)
#         if match:
#             return match.group(0).strip()
        
#         return None
    
#     def _extract_title(self, container) -> Optional[str]:
#         """Extract faculty title/position."""
#         return self._extract_title_from_text(container.get_text())
    
#     def _extract_email(self, container) -> Optional[str]:
#         """Extract email address."""
#         # First try mailto links
#         mailto_links = container.find_all('a', href=re.compile(r'^mailto:'))
#         if mailto_links:
#             href = mailto_links[0]['href']
#             return href.replace('mailto:', '')
        
#         # Then try text extraction
#         return self._extract_email_from_text(container.get_text())
    
#     def _extract_phone(self, container) -> Optional[str]:
#         """Extract phone number."""
#         text = container.get_text()
#         phone_patterns = [
#             r'\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}',
#             r'\+?1[-.\s]?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}'
#         ]
        
#         for pattern in phone_patterns:
#             match = re.search(pattern, text)
#             if match:
#                 return match.group(0).strip()
        
#         return None
    
#     def _extract_office(self, container) -> Optional[str]:
#         """Extract office location."""
#         text = container.get_text()
#         office_patterns = [
#             r'Office:?\s*([A-Z]{1,4}\s*\d+[A-Z]?)',
#             r'Room:?\s*([A-Z]{1,4}\s*\d+[A-Z]?)',
#             r'Location:?\s*([A-Z]{1,4}\s*\d+[A-Z]?)'
#         ]
        
#         for pattern in office_patterns:
#             match = re.search(pattern, text, re.IGNORECASE)
#             if match:
#                 return match.group(1).strip()
        
#         return None
    
#     def _extract_profile_url(self, container, base_url: str) -> Optional[str]:
#         """Extract profile page URL."""
#         links = container.find_all('a', href=True)
#         for link in links:
#             href = link['href']
#             full_url = urljoin(base_url, href)
            
#             # Check if link text suggests it's a profile
#             link_text = link.get_text().lower()
#             if any(word in link_text for word in ['profile', 'bio', 'cv', 'more info', 'details']):
#                 return full_url
        
#         # If no explicit profile link, return first meaningful link
#         for link in links:
#             href = link['href']
#             if href and not href.startswith('#') and not href.startswith('mailto:'):
#                 return urljoin(base_url, href)
        
#         return None
    
#     def _extract_image_url(self, container, base_url: str) -> Optional[str]:
#         """Extract faculty photo URL."""
#         img = container.find('img')
#         if img and img.get('src'):
#             return urljoin(base_url, img['src'])
#         return None
    
#     def _extract_research_areas(self, container) -> Optional[str]:
#         """Extract research areas/interests."""
#         text = container.get_text()
        
#         # Look for sections mentioning research
#         research_patterns = [
#             r'Research[:\s]+([^.]+\.)',
#             r'Interests?[:\s]+([^.]+\.)',
#             r'Areas?[:\s]+([^.]+\.)',
#             r'Specialization[:\s]+([^.]+\.)'
#         ]
        
#         for pattern in research_patterns:
#             match = re.search(pattern, text, re.IGNORECASE)
#             if match:
#                 return match.group(1).strip()
        
#         return None
    
#     def _extract_structured_data(self, soup: BeautifulSoup) -> List[Dict]:
#         """Extract faculty info from structured data (JSON-LD, microdata)."""
#         faculty_list = []
        
#         # Look for JSON-LD structured data
#         scripts = soup.find_all('script', type='application/ld+json')
#         for script in scripts:
#             try:
#                 data = json.loads(script.string)
#                 if isinstance(data, dict) and data.get('@type') == 'Person':
#                     faculty_list.append(self._parse_json_ld_person(data))
#             except:
#                 continue
        
#         return faculty_list
    
#     def _parse_json_ld_person(self, data: Dict) -> Dict:
#         """Parse JSON-LD person data."""
#         return {
#             'name': data.get('name'),
#             'title': data.get('jobTitle'),
#             'email': data.get('email'),
#             'phone': data.get('telephone'),
#             'profile_url': data.get('url'),
#             'image_url': data.get('image'),
#         }
    
#     def _remove_duplicates(self, faculty_list: List[Dict]) -> List[Dict]:
#         """Remove duplicate faculty entries."""
#         seen = set()
#         unique_faculty = []
        
#         for faculty in faculty_list:
#             # Use name and email as unique identifiers
#             name = faculty.get('name', '').lower()
#             email = faculty.get('email', '').lower()
#             identifier = (name, email)
            
#             if identifier not in seen and (name or email):
#                 seen.add(identifier)
#                 unique_faculty.append(faculty)
        
#         return unique_faculty
    
#     def scrape_faculty(self, url: str, debug=True) -> List[Dict]:
#         """Main method to scrape faculty from a URL."""
#         print(f"Scraping faculty from: {url}")
        
#         # Try both methods if Selenium is available
#         soup = self.get_page_content(url)
#         if not soup:
#             print("Failed to fetch page content")
#             return []
        
#         faculty_list = self.extract_faculty_info(soup, url, debug=debug)
        
#         print(f"Found {len(faculty_list)} faculty members")
        
#         # Show what we found
#         if faculty_list:
#             print("\nFirst few results:")
#             for i, faculty in enumerate(faculty_list[:3], 1):
#                 print(f"{i}. {faculty.get('name', 'No name')}")
#                 print(f"   Title: {faculty.get('title', 'No title')}")
#                 print(f"   Email: {faculty.get('email', 'No email')}")
#                 print()
        
#         return faculty_list
    
#     def save_to_json(self, faculty_list: List[Dict], filename: str = 'faculty.json'):
#         """Save faculty data to JSON file."""
#         with open(filename, 'w', encoding='utf-8') as f:
#             json.dump(faculty_list, f, indent=2, ensure_ascii=False)
#         print(f"Saved faculty data to {filename}")
    
#     def save_to_csv(self, faculty_list: List[Dict], filename: str = 'faculty.csv'):
#         """Save faculty data to CSV file."""
#         if not faculty_list:
#             return
        
#         fieldnames = set()
#         for faculty in faculty_list:
#             fieldnames.update(faculty.keys())
        
#         with open(filename, 'w', newline='', encoding='utf-8') as f:
#             writer = csv.DictWriter(f, fieldnames=list(fieldnames))
#             writer.writeheader()
#             writer.writerows(faculty_list)
#         print(f"Saved faculty data to {filename}")
    
#     def __del__(self):
#         """Clean up Selenium driver if it exists."""
#         if hasattr(self, 'driver'):
#             try:
#                 self.driver.quit()
#             except:
#                 pass

# # Example usage
# if __name__ == "__main__":
#     # First try with requests
#     print("=== Trying with requests (fast, but may miss JS content) ===")
#     scraper = FacultyScraper(use_selenium=False)
#     faculty_data = scraper.scrape_faculty("https://engineering.tamu.edu/cse/profiles/index.html#Faculty")
    
#     if not faculty_data and SELENIUM_AVAILABLE:
#         print("\n=== No results with requests, trying with Selenium ===")
#         scraper_selenium = FacultyScraper(use_selenium=True)
#         faculty_data = scraper_selenium.scrape_faculty("https://engineering.tamu.edu/cse/profiles/index.html#Faculty")
    
#     # Save results
#     if faculty_data:
#         scraper.save_to_json(faculty_data, 'faculty_data.json')
#         scraper.save_to_csv(faculty_data, 'faculty_data.csv')
#         print(f"\nScraping complete! Found {len(faculty_data)} total faculty members")
#     else:
#         print("\nNo faculty data found. The page might:")
#         print("1. Load content via JavaScript (try installing Selenium)")
#         print("2. Have unusual HTML structure")
#         print("3. Be blocking automated access")
#         print("4. Require authentication")
        
#         if not SELENIUM_AVAILABLE:
#             print("\nTo handle JavaScript content, install Selenium:")
#             print("pip install selenium")
#             print("And download ChromeDriver from: https://chromedriver.chromium.org/")

# scraper = FacultyScraper(use_selenium=True)  # Use Selenium for JS content
# faculty_data = scraper.scrape_faculty("https://engineering.tamu.edu/cse/profiles/index.html#Faculty", debug=True)


import requests
from bs4 import BeautifulSoup
import json
import csv
import re
from urllib.parse import urljoin, urlparse
import time
from typing import List, Dict, Optional
import pandas as pd

# Try to import Selenium for JavaScript-heavy pages
try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False
    print("Selenium not available. Install with: pip install selenium")
    print("Also download ChromeDriver from: https://chromedriver.chromium.org/")

class FacultyScraper:
    def __init__(self, use_selenium=False):
        self.use_selenium = use_selenium and SELENIUM_AVAILABLE
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })
        
        if self.use_selenium:
            self.setup_selenium_driver()
    
    def setup_selenium_driver(self):
        """Setup Selenium WebDriver with appropriate options."""
        chrome_options = Options()
        chrome_options.add_argument('--headless')  # Run in background
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1920,1080')
        
        try:
            self.driver = webdriver.Chrome(options=chrome_options)
        except Exception as e:
            print(f"Failed to setup Chrome driver: {e}")
            print("Make sure ChromeDriver is installed and in PATH")
            self.use_selenium = False
    
    def get_page_content(self, url: str) -> Optional[BeautifulSoup]:
        """Fetch and parse the webpage content."""
        if self.use_selenium:
            return self._get_content_selenium(url)
        else:
            return self._get_content_requests(url)
    
    def _get_content_requests(self, url: str) -> Optional[BeautifulSoup]:
        """Fetch content using requests."""
        try:
            response = self.session.get(url, timeout=10)
            response.raise_for_status()
            return BeautifulSoup(response.content, 'html.parser')
        except Exception as e:
            print(f"Error fetching {url} with requests: {e}")
            return None
    
    def _get_content_selenium(self, url: str) -> Optional[BeautifulSoup]:
        """Fetch content using Selenium (handles JavaScript)."""
        try:
            self.driver.get(url)
            
            # Wait for potential faculty content to load
            try:
                WebDriverWait(self.driver, 10).until(
                    lambda driver: len(driver.find_elements(By.CSS_SELECTOR, 
                        '[class*="faculty"], [class*="profile"], [class*="person"], h2, h3')) > 0
                )
            except:
                # Timeout is okay, just continue with what's loaded
                pass
            
            # Additional wait for any AJAX content
            time.sleep(3)
            
            html = self.driver.page_source
            return BeautifulSoup(html, 'html.parser')
        except Exception as e:
            print(f"Error fetching {url} with Selenium: {e}")
            return None
    
    def debug_page_structure(self, soup: BeautifulSoup, url: str):
        """Debug function to understand page structure."""
        print(f"\n=== DEBUG: Page Structure Analysis for {url} ===")
        
        # Check for common faculty-related elements
        selectors_to_check = {
            'Faculty containers': '[class*="faculty"]',
            'Profile containers': '[class*="profile"]', 
            'Person containers': '[class*="person"]',
            'Card containers': '[class*="card"]',
            'All headings': 'h1, h2, h3, h4, h5, h6',
            'Links': 'a[href]',
            'Images': 'img',
            'Email links': 'a[href*="mailto"]',
            'Divs with IDs': 'div[id]',
            'Divs with classes': 'div[class]',
        }
        
        for name, selector in selectors_to_check.items():
            elements = soup.select(selector)
            print(f"{name}: {len(elements)} found")
            
            # Show first few examples
            for i, elem in enumerate(elements[:3]):
                if name == 'All headings':
                    print(f"  {i+1}. {elem.name}: {elem.get_text()[:100].strip()}")
                elif name == 'Links':
                    print(f"  {i+1}. {elem.get('href', '')}: {elem.get_text()[:50].strip()}")
                elif name == 'Images':
                    print(f"  {i+1}. {elem.get('src', '')} - alt: {elem.get('alt', '')}")
                else:
                    classes = ' '.join(elem.get('class', []))
                    text_preview = elem.get_text()[:100].strip().replace('\n', ' ')
                    print(f"  {i+1}. class='{classes}': {text_preview}")
        
        # Check for JavaScript-loaded content indicators
        scripts = soup.find_all('script')
        js_indicators = ['ajax', 'fetch', 'XMLHttpRequest', 'faculty', 'profile']
        js_found = []
        for script in scripts:
            script_text = script.get_text().lower() if script.string else ''
            for indicator in js_indicators:
                if indicator in script_text:
                    js_found.append(indicator)
        
        if js_found:
            print(f"JavaScript indicators found: {set(js_found)}")
            print("Page likely loads content dynamically - consider using Selenium")
        
        print("=== END DEBUG ===\n")
    
    def extract_faculty_info(self, soup: BeautifulSoup, base_url: str, debug=True) -> List[Dict]:
        """Extract faculty information using multiple strategies."""
        if debug:
            self.debug_page_structure(soup, base_url)
        
        faculty_list = []
        
        # Strategy 1: Texas A&M specific patterns
        tamu_faculty = self._extract_tamu_specific(soup, base_url)
        faculty_list.extend(tamu_faculty)
        
        # Strategy 2: Common faculty card/profile containers
        faculty_containers = self._find_faculty_containers(soup)
        
        for container in faculty_containers:
            faculty_info = self._extract_from_container(container, base_url)
            if faculty_info and faculty_info not in faculty_list:
                faculty_list.append(faculty_info)
        
        # Strategy 3: Extract from headings and surrounding content
        heading_faculty = self._extract_from_headings(soup, base_url)
        faculty_list.extend(heading_faculty)
        
        # Strategy 4: Look for structured data (JSON-LD, microdata)
        structured_data = self._extract_structured_data(soup)
        faculty_list.extend(structured_data)
        
        # Remove duplicates based on name or email
        faculty_list = self._remove_duplicates(faculty_list)
        
        return faculty_list
    
    def _extract_tamu_specific(self, soup: BeautifulSoup, base_url: str) -> List[Dict]:
        """Texas A&M specific extraction patterns."""
        faculty_list = []
        
        # Look for TAMU-specific patterns
        tamu_selectors = [
            '.faculty-list .faculty-item',
            '.profile-card',
            '.person-card',
            '[data-faculty]',
            '.directory-item',
            '.profile-listing',
        ]
        
        for selector in tamu_selectors:
            elements = soup.select(selector)
            for element in elements:
                faculty_info = self._extract_from_container(element, base_url)
                if faculty_info:
                    faculty_list.append(faculty_info)
        
        return faculty_list
    
    def _extract_from_headings(self, soup: BeautifulSoup, base_url: str) -> List[Dict]:
        """Extract faculty info from headings and surrounding content."""
        faculty_list = []
        
        # Look for headings that might contain names
        headings = soup.find_all(['h2', 'h3', 'h4', 'h5'])
        
        for heading in headings:
            text = heading.get_text().strip()
            
            # Check if heading looks like a person's name
            if self._looks_like_person_name(text):
                faculty_info = {'name': text}
                
                # Look for additional info in siblings or parent
                parent = heading.parent
                if parent:
                    faculty_info.update(self._extract_from_container(parent, base_url))
                
                # Look in next siblings
                for sibling in heading.find_next_siblings():
                    sibling_text = sibling.get_text()
                    
                    # Extract email
                    email = self._extract_email_from_text(sibling_text)
                    if email and 'email' not in faculty_info:
                        faculty_info['email'] = email
                    
                    # Extract title
                    if not faculty_info.get('title'):
                        title = self._extract_title_from_text(sibling_text)
                        if title:
                            faculty_info['title'] = title
                    
                    # Stop after a few siblings or if we hit another heading
                    if sibling.name in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
                        break
                
                if len(faculty_info) > 1:  # More than just name
                    faculty_list.append(faculty_info)
        
        return faculty_list
    
    def _looks_like_person_name(self, text: str) -> bool:
        """Check if text looks like a person's name."""
        words = text.split()
        
        # Should have 2-4 words
        if len(words) < 2 or len(words) > 4:
            return False
        
        # Shouldn't contain numbers
        if any(char.isdigit() for char in text):
            return False
        
        # Shouldn't be too long
        if len(text) > 50:
            return False
        
        # Should start with capital letters
        if not all(word[0].isupper() for word in words if word):
            return False
        
        # Common non-name patterns
        exclude_patterns = ['about', 'contact', 'research', 'education', 'department']
        if any(pattern in text.lower() for pattern in exclude_patterns):
            return False
        
        return True
    
    def _extract_email_from_text(self, text: str) -> Optional[str]:
        """Extract email from text."""
        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        match = re.search(email_pattern, text)
        return match.group(0) if match else None
    
    def _extract_title_from_text(self, text: str) -> Optional[str]:
        """Extract academic title from text."""
        title_patterns = [
            r'(Professor|Prof\.)\s*[A-Za-z\s]*',
            r'(Assistant|Associate|Full)\s+Professor',
            r'(Chair|Director)\s+of\s+[A-Za-z\s]+',
            r'(Dr\.|PhD|Ph\.D\.)',
            r'(Lecturer|Instructor)',
        ]
        
        for pattern in title_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return match.group(0).strip()
        
        return None
    
    def _find_faculty_containers(self, soup: BeautifulSoup) -> List:
        """Find potential faculty containers using various selectors."""
        selectors = [
            # Common class patterns for faculty profiles
            '[class*="faculty"]',
            '[class*="profile"]',
            '[class*="person"]',
            '[class*="staff"]',
            '[class*="member"]',
            '[class*="bio"]',
            '[class*="card"]',
            '[class*="directory"]',
            # Common ID patterns
            '[id*="faculty"]',
            '[id*="profile"]',
            # Semantic elements
            'article',
            '.row > div',
            '.col > div',
            # List items that might contain faculty
            'li[class*="faculty"]',
            'li[class*="profile"]',
            'li[class*="person"]',
        ]
        
        containers = []
        for selector in selectors:
            elements = soup.select(selector)
            for element in elements:
                # Check if element likely contains faculty info
                if self._looks_like_faculty_container(element):
                    containers.append(element)
        
        return containers
    
    def _looks_like_faculty_container(self, element) -> bool:
        """Determine if an element likely contains faculty information."""
        text = element.get_text().lower()
        
        # Look for academic titles
        academic_titles = ['professor', 'dr.', 'ph.d', 'assistant', 'associate', 'chair', 'director', 'lecturer']
        has_title = any(title in text for title in academic_titles)
        
        # Look for email patterns
        has_email = '@' in text and ('edu' in text or 'ac.' in text)
        
        # Look for common faculty page indicators
        faculty_indicators = ['research', 'education', 'cv', 'publications', 'office', 'phone', 'interests']
        has_indicators = sum(1 for indicator in faculty_indicators if indicator in text) >= 1
        
        # Check for images (faculty photos)
        has_image = element.find('img') is not None
        
        # Should have reasonable amount of text
        has_content = 20 <= len(text.strip()) <= 2000
        
        # Look for person name patterns
        has_name = bool(re.search(r'\b[A-Z][a-z]+\s+[A-Z][a-z]+\b', element.get_text()))
        
        return (has_title or has_email or has_name) and has_content
    
    def _extract_from_container(self, container, base_url: str) -> Dict:
        """Extract faculty information from a container element."""
        faculty_info = {}
        
        # Extract name (usually in headings or strong/bold text)
        name = self._extract_name(container)
        if name:
            faculty_info['name'] = name
        
        # Extract title/position
        title = self._extract_title(container)
        if title:
            faculty_info['title'] = title
        
        # Extract email
        email = self._extract_email(container)
        if email:
            faculty_info['email'] = email
        
        # Extract phone
        phone = self._extract_phone(container)
        if phone:
            faculty_info['phone'] = phone
        
        # Extract office/location
        office = self._extract_office(container)
        if office:
            faculty_info['office'] = office
        
        # Extract profile URL
        profile_url = self._extract_profile_url(container, base_url)
        if profile_url:
            faculty_info['profile_url'] = profile_url
        
        # Extract image URL
        image_url = self._extract_image_url(container, base_url)
        if image_url:
            faculty_info['image_url'] = image_url
        
        # Extract research areas/interests
        research = self._extract_research_areas(container)
        if research:
            faculty_info['research_areas'] = research
        
        return faculty_info if faculty_info else {}
    
    def _extract_name(self, container) -> Optional[str]:
        """Extract faculty name from container."""
        # Try headings first
        for tag in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
            heading = container.find(tag)
            if heading:
                name = heading.get_text(strip=True)
                if self._looks_like_person_name(name):
                    return name
        
        # Try strong/bold text
        strong_tags = container.find_all(['strong', 'b'])
        for strong in strong_tags:
            text = strong.get_text(strip=True)
            if self._looks_like_person_name(text):
                return text
        
        # Try links that might be names
        links = container.find_all('a')
        for link in links:
            text = link.get_text(strip=True)
            if self._looks_like_person_name(text):
                return text
        
        # Try to find name patterns in the text
        text = container.get_text()
        name_pattern = r'\b[A-Z][a-z]+(?:\s+[A-Z][a-z]*\.?)?\s+[A-Z][a-z]+\b'
        match = re.search(name_pattern, text)
        if match:
            return match.group(0).strip()
        
        return None
    
    def _extract_title(self, container) -> Optional[str]:
        """Extract faculty title/position."""
        return self._extract_title_from_text(container.get_text())
    
    def _extract_email(self, container) -> Optional[str]:
        """Extract email address."""
        # First try mailto links
        mailto_links = container.find_all('a', href=re.compile(r'^mailto:'))
        if mailto_links:
            href = mailto_links[0]['href']
            return href.replace('mailto:', '')
        
        # Then try text extraction
        return self._extract_email_from_text(container.get_text())
    
    def _extract_phone(self, container) -> Optional[str]:
        """Extract phone number."""
        text = container.get_text()
        phone_patterns = [
            r'\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}',
            r'\+?1[-.\s]?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}'
        ]
        
        for pattern in phone_patterns:
            match = re.search(pattern, text)
            if match:
                return match.group(0).strip()
        
        return None
    
    def _extract_office(self, container) -> Optional[str]:
        """Extract office location."""
        text = container.get_text()
        office_patterns = [
            r'Office:?\s*([A-Z]{1,4}\s*\d+[A-Z]?)',
            r'Room:?\s*([A-Z]{1,4}\s*\d+[A-Z]?)',
            r'Location:?\s*([A-Z]{1,4}\s*\d+[A-Z]?)'
        ]
        
        for pattern in office_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return match.group(1).strip()
        
        return None
    
    def _extract_profile_url(self, container, base_url: str) -> Optional[str]:
        """Extract profile page URL."""
        links = container.find_all('a', href=True)
        for link in links:
            href = link['href']
            full_url = urljoin(base_url, href)
            
            # Check if link text suggests it's a profile
            link_text = link.get_text().lower()
            if any(word in link_text for word in ['profile', 'bio', 'cv', 'more info', 'details']):
                return full_url
        
        # If no explicit profile link, return first meaningful link
        for link in links:
            href = link['href']
            if href and not href.startswith('#') and not href.startswith('mailto:'):
                return urljoin(base_url, href)
        
        return None
    
    def _extract_image_url(self, container, base_url: str) -> Optional[str]:
        """Extract faculty photo URL."""
        img = container.find('img')
        if img and img.get('src'):
            return urljoin(base_url, img['src'])
        return None
    
    def _extract_research_areas(self, container) -> Optional[str]:
        """Extract research areas/interests."""
        text = container.get_text()
        
        # Look for sections mentioning research
        research_patterns = [
            r'Research[:\s]+([^.]+\.)',
            r'Interests?[:\s]+([^.]+\.)',
            r'Areas?[:\s]+([^.]+\.)',
            r'Specialization[:\s]+([^.]+\.)'
        ]
        
        for pattern in research_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return match.group(1).strip()
        
        return None
    
    def _extract_structured_data(self, soup: BeautifulSoup) -> List[Dict]:
        """Extract faculty info from structured data (JSON-LD, microdata)."""
        faculty_list = []
        
        # Look for JSON-LD structured data
        scripts = soup.find_all('script', type='application/ld+json')
        for script in scripts:
            try:
                data = json.loads(script.string)
                if isinstance(data, dict) and data.get('@type') == 'Person':
                    faculty_list.append(self._parse_json_ld_person(data))
            except:
                continue
        
        return faculty_list
    
    def _parse_json_ld_person(self, data: Dict) -> Dict:
        """Parse JSON-LD person data."""
        return {
            'name': data.get('name'),
            'title': data.get('jobTitle'),
            'email': data.get('email'),
            'phone': data.get('telephone'),
            'profile_url': data.get('url'),
            'image_url': data.get('image'),
        }
    
    def _remove_duplicates(self, faculty_list: List[Dict]) -> List[Dict]:
        """Remove duplicate faculty entries."""
        seen = set()
        unique_faculty = []
        
        for faculty in faculty_list:
            # Use name and email as unique identifiers
            name = faculty.get('name', '').lower()
            email = faculty.get('email', '').lower()
            identifier = (name, email)
            
            if identifier not in seen and (name or email):
                seen.add(identifier)
                unique_faculty.append(faculty)
        
        return unique_faculty
    
    def scrape_faculty(self, url: str, debug=True) -> List[Dict]:
        """Main method to scrape faculty from a URL."""
        print(f"Scraping faculty from: {url}")
        
        # Try both methods if Selenium is available
        soup = self.get_page_content(url)
        if not soup:
            print("Failed to fetch page content")
            return []
        
        faculty_list = self.extract_faculty_info(soup, url, debug=debug)
        
        print(f"Found {len(faculty_list)} faculty members")
        
        # Show what we found
        if faculty_list:
            print("\nFirst few results:")
            for i, faculty in enumerate(faculty_list[:3], 1):
                print(f"{i}. {faculty.get('name', 'No name')}")
                print(f"   Title: {faculty.get('title', 'No title')}")
                print(f"   Email: {faculty.get('email', 'No email')}")
                print()
        
        return faculty_list
    
    def save_to_json(self, faculty_list: List[Dict], filename: str = 'faculty.json'):
        """Save faculty data to JSON file."""
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(faculty_list, f, indent=2, ensure_ascii=False)
        print(f"Saved faculty data to {filename}")
    
    def save_to_csv(self, faculty_list: List[Dict], filename: str = 'faculty.csv'):
        """Save faculty data to CSV file."""
        if not faculty_list:
            return
        
        fieldnames = set()
        for faculty in faculty_list:
            fieldnames.update(faculty.keys())
        
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=list(fieldnames))
            writer.writeheader()
            writer.writerows(faculty_list)
        print(f"Saved faculty data to {filename}")
    
    def save_to_excel(self, faculty_list: List[Dict], filename: str = 'faculty.xlsx'):
        """Save faculty data to Excel file with nice formatting."""
        if not faculty_list:
            print("No faculty data to save to Excel")
            return
        
        # Create DataFrame
        df = pd.DataFrame(faculty_list)
        
        # Clean up the data
        df = self._clean_dataframe(df)
        
        # Create Excel writer with formatting
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Write main data
            df.to_excel(writer, sheet_name='Faculty Directory', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Faculty Directory']
            
            # Apply formatting
            self._format_excel_worksheet(workbook, worksheet, df)
            
            # Create summary sheet
            self._create_summary_sheet(writer, df, workbook)
        
        print(f"Saved faculty data to {filename}")
        print(f"Total faculty members: {len(df)}")
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and organize the DataFrame."""
        # Define preferred column order
        preferred_columns = ['name', 'title', 'email', 'phone', 'office', 'profile_url', 'image_url', 'research_areas']
        
        # Reorder columns
        available_columns = [col for col in preferred_columns if col in df.columns]
        other_columns = [col for col in df.columns if col not in preferred_columns]
        df = df[available_columns + other_columns]
        
        # Clean up data
        df = df.fillna('')  # Replace NaN with empty string
        
        # Clean names - remove extra whitespace and weird characters
        if 'name' in df.columns:
            df['name'] = df['name'].str.strip().str.replace(r'\s+', ' ', regex=True)
        
        # Clean titles
        if 'title' in df.columns:
            df['title'] = df['title'].str.strip().str.replace(r'\s+', ' ', regex=True)
        
        # Clean emails - make lowercase
        if 'email' in df.columns:
            df['email'] = df['email'].str.lower().str.strip()
        
        # Remove duplicates based on email or name
        if 'email' in df.columns and 'name' in df.columns:
            # Remove rows where both name and email are empty
            df = df[~((df['name'] == '') & (df['email'] == ''))]
            
            # Remove duplicates based on email (if email exists)
            df = df.drop_duplicates(subset=['email'], keep='first')
        
        return df
    
    def _format_excel_worksheet(self, workbook, worksheet, df):
        """Apply formatting to Excel worksheet."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Header formatting
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply header formatting
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for col in range(1, len(df.columns) + 1):
            column_letter = get_column_letter(col)
            column_name = df.columns[col-1]
            
            # Set specific widths for known columns
            if column_name == 'name':
                worksheet.column_dimensions[column_letter].width = 25
            elif column_name == 'title':
                worksheet.column_dimensions[column_letter].width = 30
            elif column_name == 'email':
                worksheet.column_dimensions[column_letter].width = 30
            elif column_name == 'research_areas':
                worksheet.column_dimensions[column_letter].width = 40
            elif column_name in ['phone', 'office']:
                worksheet.column_dimensions[column_letter].width = 15
            elif 'url' in column_name:
                worksheet.column_dimensions[column_letter].width = 35
            else:
                # Auto-adjust based on content
                max_length = max(
                    len(str(cell.value)) for cell in worksheet[column_letter]
                )
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border
        
        # Freeze the header row
        worksheet.freeze_panes = 'A2'
    
    def _create_summary_sheet(self, writer, df, workbook):
        """Create a summary sheet with statistics."""
        summary_data = {
            'Metric': [
                'Total Faculty Members',
                'Faculty with Email',
                'Faculty with Phone',
                'Faculty with Office Info',
                'Faculty with Profile URLs',
                'Faculty with Photos',
                'Faculty with Research Areas'
            ],
            'Count': [
                len(df),
                len(df[df['email'] != '']) if 'email' in df.columns else 0,
                len(df[df['phone'] != '']) if 'phone' in df.columns else 0,
                len(df[df['office'] != '']) if 'office' in df.columns else 0,
                len(df[df['profile_url'] != '']) if 'profile_url' in df.columns else 0,
                len(df[df['image_url'] != '']) if 'image_url' in df.columns else 0,
                len(df[df['research_areas'] != '']) if 'research_areas' in df.columns else 0,
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format summary sheet
        summary_worksheet = writer.sheets['Summary']
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Header formatting
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        
        for col in range(1, 3):  # Metric and Count columns
            cell = summary_worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        summary_worksheet.column_dimensions['A'].width = 25
        summary_worksheet.column_dimensions['B'].width = 10
    
    def __del__(self):
        """Clean up Selenium driver if it exists."""
        if hasattr(self, 'driver'):
            try:
                self.driver.quit()
            except:
                pass

# Example usage
if __name__ == "__main__":
    # First try with requests
    print("=== Trying with requests (fast, but may miss JS content) ===")
    scraper = FacultyScraper(use_selenium=False)
    faculty_data = scraper.scrape_faculty("https://siebelschool.illinois.edu/about/people/all-faculty")
    
    if not faculty_data and SELENIUM_AVAILABLE:
        print("\n=== No results with requests, trying with Selenium ===")
        scraper_selenium = FacultyScraper(use_selenium=True)
        faculty_data = scraper_selenium.scrape_faculty("https://siebelschool.illinois.edu/about/people/all-faculty")
    
    # Save results
    if faculty_data:
        scraper.save_to_json(faculty_data, 'faculty_data.json')
        scraper.save_to_csv(faculty_data, 'faculty_data.csv')
        scraper.save_to_excel(faculty_data, 'faculty_data.xlsx')  # Add Excel export
        print(f"\nScraping complete! Found {len(faculty_data)} total faculty members")
    else:
        print("\nNo faculty data found. The page might:")
        print("1. Load content via JavaScript (try installing Selenium)")
        print("2. Have unusual HTML structure")
        print("3. Be blocking automated access")
        print("4. Require authentication")
        
        if not SELENIUM_AVAILABLE:
            print("\nTo handle JavaScript content, install Selenium:")
            print("pip install selenium")
            print("And download ChromeDriver from: https://chromedriver.chromium.org/")
            print("\nTo export to Excel, install pandas and openpyxl:")
            print("pip install pandas openpyxl")


# Your existing code will now automatically create an Excel file
scraper = FacultyScraper(use_selenium=True)
faculty_data = scraper.scrape_faculty("https://siebelschool.illinois.edu/about/people/all-faculty")

# This will create faculty_data.xlsx automatically
scraper.save_to_excel(faculty_data, 'tamu_faculty.xlsx')